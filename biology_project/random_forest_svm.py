from random import seed
from random import randint
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import os

def split_train_test(dataset, ratio):  # 取百分之二十的数据当做测试数据
    num = len(dataset)
    print(len(dataset))
    train_num = int((1-ratio) * num)
    print(train_num)
    k = 0
    dataset_copy = list()
    while k < len(dataset):
        dataset_copy.append(dataset[k])
        k += 1
    train_data = list()
    while len(train_data) < train_num:
        index = randint(0, len(dataset_copy) - 1)
        train_data.append(dataset_copy.pop(index))

    testdata = dataset_copy
    return train_data, testdata

if __name__ == '__main__':
    seed(1)   #每一次执行本文件时都能产生同一个随机数
    datafile = load_workbook(u'D:/python_project/minder/biology_project/random_forest_data.xlsx')
    booksheet = datafile.active
    rows = booksheet.rows
    columns = booksheet.columns
    dataset = list()
    temp = list()
    i = 0
    for row in rows:
        k = 0
        i = i + 1
        line = [col.value for col in row]
        col_sum = 6
        temp = list()
        while k < col_sum:
            cell_data = booksheet.cell(row=i, column=k + 1).value
            temp.append(cell_data)
            k = k + 1
            print(temp)
        dataset.append(temp)
    print(dataset)
    print(dataset[0][:-1])
    num = 50
    x_line = [1] * num
    for k in range(num):
        x_line[k] = k + 1
    y_line = [0] * num
    for i in range(1, 51):
        ratio_ = i/100
        train_data, test_data = split_train_test(dataset, ratio=ratio_)
        with open('D:/svm/libsvm-3.24/tools/random_forest_svm_train.txt', 'w') as f:
            for k in range(len(train_data)):
                if train_data[k][5] == "RLH":
                    f.write(str("1 "))
                elif train_data[k][5] == "Lymphoma":
                    f.write(str("2 "))
                for x in range(len(train_data[k]) - 1):
                    f.write(str(x + 1) + ":" + str(train_data[k][x]) + " ")
                f.write("\n")

        with open('D:/svm/libsvm-3.24/tools/random_forest_svm_test.txt', 'w') as f:
            for k in range(len(test_data)):
                if test_data[k][5] == "RLH":
                    f.write(str("1 "))
                elif test_data[k][5] == "Lymphoma":
                    f.write(str("2 "))
                for x in range(len(test_data[k]) - 1):
                    f.write(str(x + 1) + ":" + str(test_data[k][x]) + " ")
                f.write("\n")
        try:
            cmd1 = "svm-scale -s train2.range random_forest_svm_train.txt > random_train_svm.txt"
            cmd2 = "svm-scale -r train2.range random_forest_svm_test.txt > random_test_svm.txt"
            cmd3 = "python grid.py random_train_svm.txt"

            res1 = os.popen(cmd1)
            res2 = os.popen(cmd2)
            res3 = os.popen(cmd3)
            r = res3.read()
            result = r.splitlines()
            param = result[-1].split()
            res3.close()
            cmd4 = "svm-train -c " + str(param[0]) + " -g " + str(param[1]) + " random_train_svm.txt train2.model"
            print(cmd4)
            cmd5 = "svm-predict random_test_svm.txt train2.model result2.scale"
            res4 = os.popen(cmd4)
            res5 = os.popen(cmd5)
            t = res5.read()
            result2 = t.splitlines()
            print(result2)
            acc = result2[-1].split()
            acc = str(acc[2])
            acc = acc.replace("%", "")
            y_line[i - 1] = float(acc)
            print(x_line[i - 1])
            print(y_line[i - 1])
            res5.close()
        except IOError:
            continue
    plt.title("Acc of decreasing train set", fontsize=20)
    plt.xlabel('test set ratio(%)')
    plt.ylabel('Acc(%)')
    plt.plot(x_line, y_line)
    plt.show()
    for i in range(99,  49, -1):
        ratio_ = i/100
        train_data, test_data = split_train_test(dataset, ratio=ratio_)
        with open('D:/svm/libsvm-3.24/tools/random_forest_svm_train.txt', 'w') as f:
            for k in range(len(train_data)):
                if train_data[k][5] == "RLH":
                    f.write(str("1 "))
                elif train_data[k][5] == "Lymphoma":
                    f.write(str("2 "))
                for x in range(len(train_data[k]) - 1):
                    f.write(str(x + 1) + ":" + str(train_data[k][x]) + " ")
                f.write("\n")

        with open('D:/svm/libsvm-3.24/tools/random_forest_svm_test.txt', 'w') as f:
            for k in range(len(test_data)):
                if test_data[k][5] == "RLH":
                    f.write(str("1 "))
                elif test_data[k][5] == "Lymphoma":
                    f.write(str("2 "))
                for x in range(len(test_data[k]) - 1):
                    f.write(str(x + 1) + ":" + str(test_data[k][x]) + " ")
                f.write("\n")
        try:
            cmd1 = "svm-scale -s train2.range random_forest_svm_train.txt > random_train_svm.txt"
            cmd2 = "svm-scale -r train2.range random_forest_svm_test.txt > random_test_svm.txt"
            cmd3 = "python grid.py random_train_svm.txt"

            res1 = os.popen(cmd1)
            res2 = os.popen(cmd2)
            res3 = os.popen(cmd3)
            r = res3.read()
            result = r.splitlines()
            param = result[-1].split()
            res3.close()
            cmd4 = "svm-train -c " + str(param[0]) + " -g " + str(param[1]) + " random_train_svm.txt train2.model"
            print(cmd4)
            cmd5 = "svm-predict random_test_svm.txt train2.model result2.scale"
            res4 = os.popen(cmd4)
            res5 = os.popen(cmd5)
            t = res5.read()
            result2 = t.splitlines()
            print(result2)
            acc = result2[-1].split()
            acc = str(acc[2])
            acc = acc.replace("%", "")
            y_line[i - 50] = float(acc)
            print(x_line[i - 50])
            print(y_line[i - 50])
            res5.close()
        except IOError:
            continue
    plt.plot(x_line, y_line)
    plt.show()






