from openpyxl import load_workbook

workbook = load_workbook(u'D:/python_project/minder/biology_project/data2.xlsx')

booksheet = workbook.active
rows = booksheet.rows
columns = booksheet.columns

i = 1

with open('D:/python_project/minder/biology_project/data_train.txt', 'w') as f:
    for row in rows:
        i = i + 1
        k = 0
        line = [col.value for col in row]
        cell_data_1 = booksheet.cell(row=i, column=3).value
        if cell_data_1 == "RLH":
            f.write(str("0"))
            f.write(" ")
        elif cell_data_1 == "Lymphoma":
            f.write(str("1"))
            f.write(" ")
        k = 1
        col_sum = 7
        while k < col_sum:
            cell_data = booksheet.cell(row=i, column=k+3).value
            f.write(str(k) + ":" + str(cell_data) + " ")
            k = k + 1

        f.write("\n")


